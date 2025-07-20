import datetime
import random
import os
from linebot import LineBotApi
from linebot.models import TextSendMessage
import pytz
from openpyxl import load_workbook

def main():
    # 從環境變數取得設定
    line_bot_token = os.getenv('LINE_BOT_TOKEN')
    user_id = os.getenv('LINE_USER_ID')
    
    if not line_bot_token or not user_id:
        print("❌ 錯誤：請設定環境變數 LINE_BOT_TOKEN 和 LINE_USER_ID")
        return
    
    # 初始化 LINE Bot
    line_bot_api = LineBotApi(line_bot_token)
    
    # 檢查時間範圍
    taipei_tz = pytz.timezone('Asia/Taipei')
    now = datetime.datetime.now(taipei_tz)
    current_hour = now.hour
    
    print(f"🕐 台灣時間現在是：{now.strftime('%Y-%m-%d %H:%M:%S')} ({current_hour}點)")
    
    # 檢查是否在發送時間內（8:00-17:00）
    if not (8 <= current_hour < 17):
        print(f"⏰ 不在發送時間範圍內（8:00-17:00），當前時間：{current_hour}點")
        return
    
    try:
        # 使用 openpyxl 讀取 Excel 檔案
        print("📖 正在讀取單字檔案...")
        workbook = load_workbook("vocab.xlsx")
        sheet = workbook.active
        
        # 獲取所有單字資料（跳過標題行）
        vocab_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:  # 確保英文和中文都不為空
                vocab_data.append({
                    'english': str(row[0]).strip(),
                    'chinese': str(row[1]).strip()
                })
        
        if not vocab_data:
            print("❌ 沒有找到有效的單字資料")
            return
        
        # 隨機選擇一個單字
        selected_vocab = random.choice(vocab_data)
        english_word = selected_vocab['english']
        chinese_meaning = selected_vocab['chinese']
        
        # 建立訊息
        message = f"📚 今日單字\n{english_word} : {chinese_meaning}"
        
        # 發送訊息
        line_bot_api.push_message(user_id, TextSendMessage(text=message))
        print(f"✅ 發送成功: {message}")
        
    except FileNotFoundError:
        print("❌ 找不到 vocab.xlsx 檔案")
    except Exception as e:
        print(f"❌ 發送失敗：{e}")

if __name__ == '__main__':  # ✅ 正確：雙底線
main()
